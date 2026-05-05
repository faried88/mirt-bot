import logging
import os
import math
import time
import tempfile
import json
from datetime import datetime
from io import BytesIO

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ContextTypes, ConversationHandler, filters
)

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# ── Config ───────────────────────────────────────────────────────────────────
TOKEN         = "8687940667:AAE2jfP57jzaRu5LE0BneU3IlXCcrTwJIeQ"
ALLOWED_USER  = 910169628
DRIVE_FOLDER  = "1RmxvQKfY8e9-BL3gphaVNwRApNQa2YOp"
MIRT_URL      = "https://sps-bs.cigalah.com.sa"
MIRT_USER     = "M.faried88"
MIRT_PASS     = "handle"

SERVICE_ACCOUNT_INFO = {
    "type": "service_account",
    "project_id": "mirt-495420",
    "private_key_id": "eea07b7b52c951a0003fa1664127ea81f271dd4f",
    "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQCe5s79PXfwcCyK\n+z0ljxAMwDTRfcEZO3CqF7RA057THK1k72veCpIUkmle1oCcCFe6LTg0+9pVcGae\nhKAmK0e8YwBu6W/UEwAFyYB0ickvr5hcwzg870V0Lnk4dwupA34mlBqhx6JCaY4C\nR3PJ1vzVGwoUbhfU5M8Be7TJwor1DVrNugHZm7zVGnnrtJoyq0a8CZLD3nOHdzul\ngj4o8xnOTBZVN55eZmoVHnHwmO8wDG53cpnX2keESsZTV/DdLAUp5JkY/f7HLBjg\njc4s5cAt8giAC1QREiSN4pXhCCe97fecbNaSX056u46ByOnbzkkJV2uPRsAtrprU\nK4TJGHTvAgMBAAECggEAOIArY4JizlykifqRoHRBKbeCUGcdrSIkimaJUm+szrYo\ntXYobbhmfugcjXtKGbEhuHJxxO00kiK4am8QHuJOzJ6LPeTFPaxP2r7ubQG9RrZy\nP7GuooQVtxz7P2ec/sjeJ0uMOLAqcuDjfM35TvCh0AigSelnkeyV6poZC5CgJke8\n7cS0gXsUjeZz7oZra5PQlv54ymXxHx0vXhH6MAOhCV2Qks3BUdKeyy3sL/YZ48HV\nVETfxCRRKrGnFtjg1i/h10ntev7qDBW0FdF0N8y+Ju4Pgm7tGOnvc7HnTM6pESF6\nEoRfzNJtdihdygMQRHT17qIYqAb+LPZxaxswNfZTsQKBgQDKw3MM+XFQSSg3SGHH\nh8fqGXCYxeGT7vfBKenL1gWT0eWr5CfONllOC9uHA/ChEci9D0Y6O5uWDQK+YdKu\ndBhOJww2S8V6KWkhmfse/jGAa9zgGAnYeVPufE4nK6QXRPYagLChbp2hHJt5PTlF\n7UIqIMo9l8TWk9QXuKj7crQUJwKBgQDInzm3O5flbwKIMFI9xGqVR5G9Q7pfB4FU\nah0KkGJHaVEX26cKgZXdJSsQTJIMu5rUqLCvWv0mKsHEHLd6g4XP6fF1NuvCqd7e\nJUfFuOpuXPT0Ml5nswrJ392WG0MFTWdmPjbkev3XxB5ZEUU7J8fHpHNklh99QqoH\nK/AX7iot+QKBgBsWPyFllVinXUL9XWqdXfyNB3ixPrBXhSt94OjFH5uet7Ld2N94\nbTe658nCofuyd4GiL7yJyAAkntA2G0II6lJObxg1yRzHuW6utlhulshUIH6jV3Ve\nx/KdEoezEcm2AbaKqI34TACA5NgucJ9B0cv082+E/du4heXhWlm0+g+TAoGBAIOd\nyiyGoSE5Ec0s/ldda5shx+AF9dfwQY2SzBipHoDA/B2N0emXmCzr/HOF+G74CRyo\nyrlQFTIb7ODvAgQTEw+S6ADBFiywavEMPijeJpZez6kA/mRD1rkX7/RRUEfDPymZ\neUOt2KjcFhjStruXXn6ASd/ciS4RNSDdV3crnWppAoGBAMawFgwhlAylTCUs6YTH\nrRRrYtb+3dD1NX9QqlpK47oy1tq65U/8JViQ3UaA08KTwqNnXSNaxx6DjeuGk3l4\n+XAfV50ubEUW109RRChAXkrUXmm15A5wBY1TwfuivK76BKpl+rRE1fVCYqmf0CBH\nPfmpTBj1LBBlqmE1ekE9cp1X\n-----END PRIVATE KEY-----\n",
    "client_email": "mirt-bot@mirt-495420.iam.gserviceaccount.com",
    "client_id": "115674558839677734201",
    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
    "token_uri": "https://oauth2.googleapis.com/token",
    "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
    "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/mirt-bot%40mirt-495420.iam.gserviceaccount.com",
    "universe_domain": "googleapis.com"
}

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

WAIT_DATE = 0

# ════════════════════════════════════════════════════════════════════════════
# GOOGLE DRIVE
# ════════════════════════════════════════════════════════════════════════════

def get_drive_service():
    creds = Credentials.from_service_account_info(
        SERVICE_ACCOUNT_INFO,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)

def upload_to_drive(file_bytes: bytes, filename: str, folder_id: str) -> str:
    service = get_drive_service()
    meta = {"name": filename, "parents": [folder_id]}
    media = MediaIoBaseUpload(
        BytesIO(file_bytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True
    )
    f = service.files().create(body=meta, media_body=media, fields="id,webViewLink").execute()
    return f.get("webViewLink", "")

def create_drive_subfolder(name: str, parent_id: str) -> str:
    service = get_drive_service()
    meta = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id]
    }
    f = service.files().create(body=meta, fields="id").execute()
    return f.get("id")

# ════════════════════════════════════════════════════════════════════════════
# MIRT DOWNLOADER
# ════════════════════════════════════════════════════════════════════════════

REPORT_CONFIG = {
    "attendance": {"label": "📋 Attendance", "menu_text": "Attendance Report", "wait1": 120, "wait2": 0,  "part_header": False},
    "f2f":        {"label": "🤝 F2F",        "menu_text": "Master F2F",        "wait1": 120, "wait2": 120, "part_header": True},
    "market":     {"label": "📊 Market",     "menu_text": "Market Intelligence","wait1": 60,  "wait2": 0,  "part_header": False},
    "sv":         {"label": "👔 S.V",        "menu_text": "S.V Visit",          "wait1": 60,  "wait2": 0,  "part_header": False},
    "oos":        {"label": "📦 OOS",        "menu_text": "OOS Report",         "wait1": 60,  "wait2": 0,  "part_header": False},
    "callcenter": {"label": "📞 Callcenter", "menu_text": "Call Center",        "wait1": 60,  "wait2": 0,  "part_header": False},
}

def download_from_mirt(report_key: str, date_str: str, download_dir: str) -> str:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })

    driver = webdriver.Chrome(options=opts)
    wait   = WebDriverWait(driver, 30)
    cfg    = REPORT_CONFIG[report_key]

    try:
        # ── Login ────────────────────────────────────────────────────────
        driver.get(MIRT_URL)
        wait.until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(MIRT_USER)
        driver.find_element(By.NAME, "password").send_keys(MIRT_PASS)
        driver.find_element(By.CSS_SELECTOR, "button[type=submit], input[type=submit]").click()
        time.sleep(4)

        # ── Navigate ──────────────────────────────────────────────────────
        nav = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//*[contains(text(),'{cfg['menu_text']}')]")
        ))
        nav.click()
        time.sleep(3)

        # ── Set date ──────────────────────────────────────────────────────
        for selector in [
            "input[type='date']",
            "input[name*='date']",
            "input[id*='date']",
            "input[placeholder*='date']",
            "input[placeholder*='Date']",
        ]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, selector)
                driver.execute_script("arguments[0].value = arguments[1]", el, date_str)
                break
            except Exception:
                continue

        # ── Click Show ────────────────────────────────────────────────────
        show = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//button[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'show')] | //input[contains(translate(@value,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'show')]")
        ))
        show.click()
        time.sleep(cfg["wait1"])

        # ── Part header (F2F only) ────────────────────────────────────────
        if cfg["part_header"]:
            try:
                ph = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[contains(text(),'Show Report Part Header')]")
                ))
                ph.click()
                time.sleep(cfg["wait2"])
            except Exception:
                pass

        # ── Export Excel ──────────────────────────────────────────────────
        before = set(os.listdir(download_dir))
        try:
            exp = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'excel')] | //*[contains(@class,'excel')] | //*[contains(@title,'Excel')]")
            ))
            exp.click()
        except Exception:
            # Try any export/download button
            for kw in ["Export", "Download", "تحميل", "تصدير"]:
                try:
                    driver.find_element(By.XPATH, f"//*[contains(text(),'{kw}')]").click()
                    break
                except Exception:
                    continue

        # ── Wait for file ─────────────────────────────────────────────────
        for _ in range(90):
            time.sleep(2)
            after    = set(os.listdir(download_dir))
            new_xlsx = [f for f in (after - before)
                        if (f.endswith(".xlsx") or f.endswith(".xls"))
                        and not f.endswith(".crdownload")]
            if new_xlsx:
                return os.path.join(download_dir, new_xlsx[0])

        return None

    finally:
        driver.quit()

# ════════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ════════════════════════════════════════════════════════════════════════════

def set_header_style(ws, header_row=1, bg="1F4E79"):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center")

def auto_col_width(ws):
    for col in ws.columns:
        ml = 0
        cl = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value: ml = max(ml, len(str(cell.value)))
            except Exception: pass
        ws.column_dimensions[cl].width = min(ml + 4, 40)

def col_fmt_date(ws, cl, start_row=2):
    for cell in ws[cl][start_row-1:]:
        cell.number_format = "DD-MM-YY"

def col_fmt_number(ws, cl, start_row=2):
    for cell in ws[cl][start_row-1:]:
        if cell.value not in (None,""):
            try: cell.value = float(cell.value); cell.number_format = "0"
            except Exception: pass

def col_fmt_integer(ws, cl, start_row=2):
    for cell in ws[cl][start_row-1:]:
        if cell.value not in (None,""):
            try: cell.value = int(float(str(cell.value).replace(" ",""))); cell.number_format = "0"
            except Exception: pass

def df_to_sheet(wb, df, sheet_name):
    if sheet_name in wb.sheetnames: del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for c, col in enumerate(df.columns, 1): ws.cell(1, c, col)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1): ws.cell(r, c, val)
    set_header_style(ws); auto_col_width(ws)
    return ws

# ════════════════════════════════════════════════════════════════════════════
# PROCESSORS
# ════════════════════════════════════════════════════════════════════════════

def process_attendance(fb: bytes, day: str) -> bytes:
    df = pd.read_excel(BytesIO(fb), header=4)
    df.columns = df.columns.str.strip()
    df = df[df["City"].notna() & (df["City"].astype(str).str.strip() != "") & (df["City"].astype(str).str.strip() != "City")].reset_index(drop=True)
    df["Project Type"] = df.groupby("Tablet #")["Project Type"].transform(lambda x: x.ffill().bfill())
    non_agg = {c: (c,"first") for c in df.columns if c not in ("Tablet #","Check In","Check Out")}
    agg = df.groupby("Tablet #", as_index=False).agg(**non_agg, **{"Check In":("Check In","min"),"Check Out":("Check Out","max")})
    try: agg = agg[df.columns]
    except Exception: pass
    agg = agg.sort_values("Check In").drop_duplicates("Tablet #", keep="last").reset_index(drop=True)
    wb = load_workbook(BytesIO(fb))
    for s in wb.sheetnames: del wb[s]
    ws = df_to_sheet(wb, agg, "Detailed Attendance")
    cols = list(agg.columns)
    if "Date"        in cols: col_fmt_date(ws,    get_column_letter(cols.index("Date")+1))
    if "Outlet Code" in cols: col_fmt_number(ws,  get_column_letter(cols.index("Outlet Code")+1))
    if "Mobile"      in cols: col_fmt_integer(ws, get_column_letter(cols.index("Mobile")+1))
    out = BytesIO(); wb.save(out); return out.getvalue()

def process_callcenter(fb: bytes, day: str) -> bytes:
    df = pd.read_excel(BytesIO(fb), header=8).dropna(axis=1,how="all").dropna(how="all")
    df.columns = df.columns.str.strip()
    wb = load_workbook(BytesIO(fb))
    for s in wb.sheetnames: del wb[s]
    df_to_sheet(wb, df, "Callcenter Report")
    out = BytesIO(); wb.save(out); return out.getvalue()

def _f2f_qr(row):
    p = str(row.get("Project","")).strip()
    op = str(row.get("Our Products","")).strip()
    if p == "Davidoff":
        if any(k in op for k in ["DVD Gold","DVD Slim Gold","DVD Slim White","DVD White"]): return "DVD PL"
        if "Evolve" in op: return "Evolve"
    return p

def process_f2f(fb: bytes, day: str) -> bytes:
    df = pd.read_excel(BytesIO(fb), header=0)
    df.columns = df.columns.str.strip()
    for col in df.columns:
        df = df[~(df[col].astype(str).str.strip().str.lower() == "grand total")]
    df = df.reset_index(drop=True)
    if "Status" in df.columns and "Mobile" in df.columns:
        df.loc[df["Status"].astype(str).str.strip() == "Refused to flip","Mobile"] = "0000000000"
    df["Project QR"] = df.apply(_f2f_qr, axis=1)
    wb = load_workbook(BytesIO(fb))
    for s in wb.sheetnames: del wb[s]
    ws_n = df_to_sheet(wb, df, "Numbers")
    if "Mobile" in df.columns and "Status" in df.columns:
        al = df[df["Status"].astype(str).str.strip().isin(["Agreed","Loyal"])]
        dv = set(al[al.duplicated("Mobile",keep=False)]["Mobile"].astype(str))
        mi = list(df.columns).index("Mobile")+1
        yf = PatternFill("solid",fgColor="FFFF00")
        for r in range(2, ws_n.max_row+1):
            if str(ws_n.cell(r,mi).value) in dv: ws_n.cell(r,mi).fill = yf
    df_to_sheet(wb, df, "Main")
    if "QR Code" in df.columns:
        hq = df[df["QR Code"].notna() & (df["QR Code"].astype(str).str.strip()!="")]
        nq = df[df["QR Code"].isna()  | (df["QR Code"].astype(str).str.strip()=="")]
        qdf = pd.concat([hq,nq],ignore_index=True)
        ws_q = df_to_sheet(wb, qdf, "QR")
        qi = list(qdf.columns).index("QR Code")+1
        al2 = qdf[qdf["Status"].astype(str).str.strip().isin(["Agreed","Loyal"])] if "Status" in qdf.columns else qdf
        dq = set(al2[al2.duplicated("QR Code",keep=False)]["QR Code"].astype(str)) if not al2.empty else set()
        yf2 = PatternFill("solid",fgColor="FFFF00")
        of  = PatternFill("solid",fgColor="FFA500")
        for r in range(2, ws_q.max_row+1):
            c = ws_q.cell(r,qi)
            if c.value in (None,""): c.fill = yf2
            elif str(c.value) in dq: c.fill = of
    out = BytesIO(); wb.save(out); return out.getvalue()

def _strip_ms(fb):
    dr = pd.read_excel(BytesIO(fb), header=None)
    ms = next((i for i,row in dr.iterrows() if row.astype(str).str.upper().str.contains("MS CHOICE").any()), None)
    if ms is not None: dr = dr.drop(index=ms).reset_index(drop=True)
    df = dr.copy(); df.columns = df.iloc[0].astype(str).str.strip()
    df = df.iloc[1:].reset_index(drop=True)
    return df.loc[:, df.columns.notna() & (df.columns!="nan")]

def _pt(row):
    pr = str(row.get("Project","")).strip().upper()
    pq = str(row.get("Project Type?","")).strip()
    pt = str(row.get("Project Type","")).strip()
    pqv = pq if "GITANES" in pr else ("DVD" if "DAVIDOFF" in pr else pq)
    ptf = (pqv if not pt or pt=="nan" else ("Davidoff" if pt.upper()=="DVD" else pt))
    return pqv, ptf

def process_market(fb: bytes, day: str) -> bytes:
    df = _strip_ms(fb)
    if "Outlet ID" in df.columns:
        df = df[df["Outlet ID"].notna() & (df["Outlet ID"].astype(str).str.strip()!="")]
    if "Date" in df.columns: df["Date"] = df["Date"].ffill().bfill()
    if "MS Report Time" in df.columns: df["MS Report Time"] = df["MS Report Time"].ffill().bfill()
    if "Project Type?" in df.columns and "Project Type" in df.columns:
        r = df.apply(_pt,axis=1); df["Project Type?"]=[x[0] for x in r]; df["Project Type"]=[x[1] for x in r]
    cols = list(df.columns)
    ordered = [c for c in ["Date","MS Report Time"] if c in cols]
    df = df[ordered + [c for c in cols if c not in ordered]]
    wb = load_workbook(BytesIO(fb))
    for s in wb.sheetnames: del wb[s]
    ws = df_to_sheet(wb, df, "Market Report")
    cols = list(df.columns)
    if "Date"        in cols: col_fmt_date(ws,    get_column_letter(cols.index("Date")+1))
    if "Outlet Code" in cols: col_fmt_number(ws,  get_column_letter(cols.index("Outlet Code")+1))
    if "Mobile"      in cols: col_fmt_integer(ws, get_column_letter(cols.index("Mobile")+1))
    out = BytesIO(); wb.save(out); return out.getvalue()

FS = {"Ihab Kader","Mohammed Faried","Ahmad Gamal","Ibrahim Fawaz"}
CT = {"Mohammed Izzat","Mohamed Abdul Aziz"}

def process_sv(fb: bytes, day: str) -> bytes:
    df = _strip_ms(fb)
    if "Date" in df.columns: df["Date"] = df["Date"].ffill().bfill()
    if "MS Report Time" in df.columns: df["MS Report Time"] = df["MS Report Time"].ffill().bfill()
    if "Project Type?" in df.columns and "Project Type" in df.columns:
        r = df.apply(_pt,axis=1); df["Project Type?"]=[x[0] for x in r]; df["Project Type"]=[x[1] for x in r]
    sv_col  = next((c for c in df.columns if "s.v" in c.lower() or "coacher" in c.lower()), None)
    sup_col = next((c for c in df.columns if "supervisor" in c.lower()), None)
    if sv_col:
        def cl(n):
            n=str(n).strip()
            return ("Field Supervisor",n) if n in FS else (("Control",n) if n in CT else ("S.V",n))
        r2=df[sv_col].apply(cl); df["Type"]=[x[0] for x in r2]; df["Coacher Name"]=[x[1] for x in r2]
    wb = load_workbook(BytesIO(fb))
    for s in wb.sheetnames: del wb[s]
    ws = df_to_sheet(wb, df, "S.V Report")
    if sup_col and "Coacher Name" in list(df.columns):
        cc=ws.max_column+1
        hc=ws.cell(1,cc,"Compare"); hc.fill=PatternFill("solid",fgColor="1F4E79"); hc.font=Font(bold=True,color="FFFFFF")
        sl=get_column_letter(list(df.columns).index(sup_col)+1)
        cll=get_column_letter(list(df.columns).index("Coacher Name")+1)
        for r in range(2,ws.max_row+1): ws.cell(r,cc,f"={sl}{r}={cll}{r}")
    if "Date" in list(df.columns): col_fmt_date(ws, get_column_letter(list(df.columns).index("Date")+1))
    out = BytesIO(); wb.save(out); return out.getvalue()

def process_oos(fb: bytes, day: str) -> bytes:
    df = _strip_ms(fb)
    for col in ["MS Report Time","Promoter ID"]:
        if col in df.columns: df=df.drop(columns=[col])
    if "Outlet ID" in df.columns:
        df=df[df["Outlet ID"].notna() & (df["Outlet ID"].astype(str).str.strip()!="")]
    if "Date" in df.columns: df["Date"]=df["Date"].ffill().bfill()
    if "Project Type?" in df.columns and "Project Type" in df.columns:
        r=df.apply(_pt,axis=1); df["Project Type?"]=[x[0] for x in r]; df["Project Type"]=[x[1] for x in r]
    wb = load_workbook(BytesIO(fb))
    for s in wb.sheetnames: del wb[s]
    ws = df_to_sheet(wb, df, "OOS Report")
    cols = list(df.columns)
    if "Date"        in cols: col_fmt_date(ws, get_column_letter(cols.index("Date")+1))
    if "Outlet Code" in cols: col_fmt_number(ws, get_column_letter(cols.index("Outlet Code")+1))
    for col in cols:
        if any(k in col.lower() for k in ["qty","quantity","count"]):
            col_fmt_number(ws, get_column_letter(cols.index(col)+1))
    out = BytesIO(); wb.save(out); return out.getvalue()

PROCESSORS = {
    "attendance": (process_attendance, "attendance report"),
    "callcenter": (process_callcenter, "Callcenter Report"),
    "f2f":        (process_f2f,        "f2f report"),
    "market":     (process_market,     "Market Report"),
    "sv":         (process_sv,         "S.V Report"),
    "oos":        (process_oos,        "OOS Report"),
}

# ════════════════════════════════════════════════════════════════════════════
# BOT HANDLERS
# ════════════════════════════════════════════════════════════════════════════

def is_allowed(update: Update) -> bool:
    return update.effective_user.id == ALLOWED_USER

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update):
        await update.message.reply_text("❌ مش مصرح ليك.")
        return ConversationHandler.END
    await update.message.reply_text(
        "👋 *أهلاً يا محمد!*\n\n"
        "📅 اكتب التاريخ اللي عايز تنزل تقاريره:\n"
        "الصيغة: `YYYY-MM-DD`\n"
        "مثال: `2025-01-20`",
        parse_mode="Markdown"
    )
    return WAIT_DATE

async def handle_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return ConversationHandler.END

    date_text = update.message.text.strip()
    try:
        date_obj = datetime.strptime(date_text, "%Y-%m-%d")
    except ValueError:
        await update.message.reply_text(
            "⚠️ صيغة التاريخ غلط!\nاكتب زي كده: `2025-01-20`",
            parse_mode="Markdown"
        )
        return WAIT_DATE

    day          = str(date_obj.day)
    date_display = date_obj.strftime("%d/%m/%Y")

    msg = await update.message.reply_text(
        f"✅ تاريخ: *{date_display}*\n\n⏳ بدأت تنزيل الـ 6 تقارير من MiRT...\n🔄 استنى شوية",
        parse_mode="Markdown"
    )

    # Create dated subfolder in Drive
    folder_name = f"تقارير {date_display}"
    try:
        sub_folder_id = create_drive_subfolder(folder_name, DRIVE_FOLDER)
    except Exception as e:
        logger.warning(f"Subfolder error: {e}")
        sub_folder_id = DRIVE_FOLDER

    results = {}
    done    = 0
    total   = len(PROCESSORS)

    with tempfile.TemporaryDirectory() as tmp_dir:
        for key, (processor, fname_prefix) in PROCESSORS.items():
            label = REPORT_CONFIG[key]["label"]
            await msg.edit_text(
                f"🔄 *{label}* — جاري التنزيل...\n"
                f"({done}/{total} خلصوا)",
                parse_mode="Markdown"
            )
            try:
                file_path = download_from_mirt(key, date_text, tmp_dir)
                if not file_path or not os.path.exists(file_path):
                    raise Exception("الملف مش اتنزل")

                with open(file_path, "rb") as f:
                    file_bytes = f.read()

                await msg.edit_text(
                    f"⚙️ *{label}* — جاري المعالجة...\n"
                    f"({done}/{total} خلصوا)",
                    parse_mode="Markdown"
                )

                result_bytes = processor(file_bytes, day)
                out_name     = f"{fname_prefix} {day}.xlsx"
                drive_link   = upload_to_drive(result_bytes, out_name, sub_folder_id)
                results[key] = ("✅", label, drive_link, result_bytes, out_name)

            except Exception as e:
                logger.exception(f"Error on {key}")
                results[key] = ("❌", label, str(e), None, None)

            done += 1

    # ── Summary ───────────────────────────────────────────────────────────
    lines   = [f"📊 *تقارير {date_display}*\n"]
    all_ok  = True
    for key, data in results.items():
        status, label = data[0], data[1]
        if status == "✅":
            lines.append(f"✅ {label}")
        else:
            lines.append(f"❌ {label}\n   _{data[2][:60]}_")
            all_ok = False

    folder_link = f"https://drive.google.com/drive/folders/{sub_folder_id}"
    if all_ok:
        lines.append(f"\n🎉 *كل التقارير جاهزة!*\n📁 [افتح الفولدر على Drive]({folder_link})")
    else:
        lines.append(f"\n⚠️ بعض التقارير فيها مشكلة\n📁 [افتح الفولدر]({folder_link})")

    await msg.edit_text("\n".join(lines), parse_mode="Markdown")

    # Send each file directly in Telegram too
    for key, data in results.items():
        status, label, _, result_bytes, out_name = data
        if status == "✅" and result_bytes:
            await update.message.reply_document(
                document=BytesIO(result_bytes),
                filename=out_name,
                caption=f"📎 {label}"
            )

    # Offer next action
    keyboard = [
        [InlineKeyboardButton("📅 تاريخ تاني", callback_data="again")],
        [InlineKeyboardButton("🚪 خروج",        callback_data="exit")]
    ]
    await update.message.reply_text(
        "عايز تشتغل على تاريخ تاني؟",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return WAIT_DATE

async def callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "again":
        await query.edit_message_text(
            "📅 اكتب التاريخ:\n`YYYY-MM-DD`\nمثال: `2025-01-20`",
            parse_mode="Markdown"
        )
        return WAIT_DATE
    await query.edit_message_text("👋 تمام! لما تحتاجني ابعت /start")
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ تم الإلغاء. ابعت /start تاني.")
    return ConversationHandler.END

# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    app = Application.builder().token(TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            WAIT_DATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, handle_date),
                CallbackQueryHandler(callback_handler),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(conv)
    logger.info("✅ MiRT Bot started!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
